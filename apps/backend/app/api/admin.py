"""
Admin-only endpoints for user management and bulk import.

All routes require an active admin account (role == 'admin').
"""
from typing import Any, List, Optional
import io

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_

from app.db.database import get_db
from app.models.user import User
from app.schemas.user import UserResponse
from app.schemas.admin import AdminUserUpdate, AdminUserListResponse, BulkImportResult
from app.api.auth import get_current_user
from app.utils.password import hash_password
from app.utils.phone import normalize_phone_number

router = APIRouter(prefix="/api/admin", tags=["admin"])


# ── Dependency: admin-only guard ───────────────────────────────────────────

def get_admin_user(current_user: User = Depends(get_current_user)) -> User:
    """Raise 403 unless the authenticated user is an admin."""
    if current_user.role != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Admin access required",
        )
    return current_user


# ── Endpoints ──────────────────────────────────────────────────────────────

@router.get("/users", response_model=AdminUserListResponse)
def list_users(
    search: Optional[str] = Query(None, description="Search by name, email, IC or phone"),
    role: Optional[str] = Query(None, description="Filter by role"),
    is_active: Optional[bool] = Query(None, description="Filter by active status"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _admin: User = Depends(get_admin_user),
) -> Any:
    """List all users with optional search and filters (paginated)."""
    q = db.query(User)

    if search:
        term = f"%{search}%"
        q = q.filter(
            or_(
                User.full_name.ilike(term),
                User.email.ilike(term),
                User.ic_number.ilike(term),
                User.passport_number.ilike(term),
                User.phone_number.ilike(term),
            )
        )

    if role:
        q = q.filter(User.role == role)

    if is_active is not None:
        q = q.filter(User.is_active == is_active)

    total = q.count()
    users = q.order_by(User.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    return AdminUserListResponse(
        total=total,
        page=page,
        page_size=page_size,
        users=users,
    )


@router.get("/users/export")
def export_users(
    db: Session = Depends(get_db),
    _admin: User = Depends(get_admin_user),
):
    """Export all users to an Excel file."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed on server. Contact administrator.",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    headers = [
        "ID", "Email", "Full Name", "Phone Number", "IC Number", "Passport Number",
        "Date of Birth", "Place of Birth", "Sex", "Race", "Marital Status", 
        "Num Dependents", "Taman Name", "House Number", "Jalan", 
        "Job Title", "Employer Name", "Employer Address", "Employer Phone", 
        "Resident Type", "Committee Title", "Role", "Is Active", "Status", "Created At"
    ]
    ws.append(headers)

    users = db.query(User).order_by(User.created_at.desc()).all()
    for u in users:
        row = [
            str(u.id), u.email, u.full_name, u.phone_number, u.ic_number, u.passport_number,
            u.date_of_birth.isoformat() if u.date_of_birth else None,
            u.place_of_birth, u.sex, u.race, u.marital_status,
            u.num_dependents, u.taman_name, u.house_number, u.jalan_aman_serenia,
            u.job_title, u.employer_name, u.employer_address, u.employer_phone,
            u.resident_type, u.committee_title, u.role, 
            str(u.is_active), u.status, 
            u.created_at.isoformat() if u.created_at else None
        ]
        ws.append(row)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    headers = {
        'Content-Disposition': 'attachment; filename="ra_users_export.xlsx"'
    }
    return StreamingResponse(
        stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers=headers
    )

@router.get("/users/template")
def download_template(
    _admin: User = Depends(get_admin_user),
):
    """Generate an Excel template for importing users."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed on server. Contact administrator.",
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Template"

    headers = [
        "email", "full_name", "password", "phone_number", "ic_number", "passport_number",
        "date_of_birth", "place_of_birth", "sex", "race", "marital_status",
        "num_dependents", "taman_name", "house_number", "jalan_aman_serenia",
        "job_title", "employer_name", "employer_address", "employer_phone",
        "resident_type", "committee_title", "role", "is_active"
    ]
    ws.append(headers)

    example = [
        "resident@example.com", "Ahmad bin Ibrahim", "ChangeMe123!", "+60 12-345 6789",
        "900101-14-5678", "", "1990-01-01", "Kuala Lumpur", "M", "Malay", "married",
        2, "Serenia Amani", "12A", "Jalan Aman 3",
        "Engineer", "Acme Sdn Bhd", "No 1 Jalan Tech", "+60 3-1234 5678",
        "owner", "", "resident", "true"
    ]
    ws.append(example)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    res_headers = {
        'Content-Disposition': 'attachment; filename="ra_users_import_template.xlsx"'
    }
    return StreamingResponse(
        stream, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
        headers=res_headers
    )


@router.get("/users/{user_id}", response_model=UserResponse)
def get_user(
    user_id: str,
    db: Session = Depends(get_db),
    _admin: User = Depends(get_admin_user),
) -> Any:
    """Get a single user by ID."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    return user


@router.patch("/users/{user_id}", response_model=UserResponse)
def update_user(
    user_id: str,
    update_data: AdminUserUpdate,
    db: Session = Depends(get_db),
    _admin: User = Depends(get_admin_user),
) -> Any:
    """Admin can update role, is_active, is_verified and any profile field."""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    for field, value in update_data.model_dump(exclude_unset=True).items():
        if field == "password" and value:
            setattr(user, "password_hash", hash_password(value))
        else:
            if field == "phone_number":
                value = normalize_phone_number(value)
            setattr(user, field, value)

    db.add(user)
    db.commit()
    db.refresh(user)
    return user


@router.delete("/users/{user_id}", status_code=204)
def delete_user(
    user_id: str,
    db: Session = Depends(get_db),
    admin: User = Depends(get_admin_user),
) -> None:
    """Permanently delete a user. Admin cannot delete themselves."""
    if str(admin.id) == user_id:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    db.delete(user)
    db.commit()


@router.post("/users/import", response_model=BulkImportResult)
def import_users_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _admin: User = Depends(get_admin_user),
) -> Any:
    """
    Bulk-import users from an Excel (.xlsx) file.

    Expected columns (case-insensitive, order does not matter):
      email*, full_name*, password*, phone_number, ic_number, passport_number,
      date_of_birth (YYYY-MM-DD), place_of_birth, sex, race,
      marital_status, num_dependents, taman_name, house_number,
      jalan_aman_serenia, job_title, employer_name,
      employer_address, employer_phone, role, is_active

    Rows with missing email or full_name are skipped.
    Duplicate emails (already in DB) are skipped.
    """
    try:
        import openpyxl  # type: ignore
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed on server. Contact administrator.",
        )

    if not file.filename or not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    contents = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Could not parse Excel file")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    # Normalise header names
    headers = [str(h).strip().lower().replace(" ", "_") if h else "" for h in rows[0]]

    def col(name: str) -> Optional[int]:
        try:
            return headers.index(name)
        except ValueError:
            return None

    created, skipped, errors = 0, 0, []

    for row_num, row in enumerate(rows[1:], start=2):
        def val(name: str) -> Optional[str]:
            idx = col(name)
            if idx is None or idx >= len(row):
                return None
            v = row[idx]
            return str(v).strip() if v is not None else None

        email = val("email")
        full_name = val("full_name")

        if not email or not full_name:
            skipped += 1
            continue

        # Skip duplicates
        if db.query(User).filter(User.email == email).first():
            skipped += 1
            errors.append(f"Row {row_num}: {email} already exists — skipped")
            continue

        raw_pwd = val("password") or "ChangeMe123!"
        from datetime import date as _date

        dob = None
        dob_str = val("date_of_birth")
        if dob_str:
            try:
                dob = _date.fromisoformat(dob_str)
            except ValueError:
                pass

        num_dep = None
        nd_str = val("num_dependents")
        if nd_str:
            try:
                num_dep = int(float(nd_str))
            except ValueError:
                pass

        role_val = val("role") or "resident"
        if role_val not in ("admin", "resident"):
            role_val = "resident"

        is_active_val = val("is_active")
        active = is_active_val.lower() in ("true", "1", "yes", "y") if is_active_val else True

        user = User(
            email=email,
            password_hash=hash_password(raw_pwd),
            full_name=full_name,
            phone_number=normalize_phone_number(val("phone_number")),
            ic_number=val("ic_number"),
            passport_number=val("passport_number"),
            date_of_birth=dob,
            place_of_birth=val("place_of_birth"),
            sex=val("sex"),
            race=val("race"),
            marital_status=val("marital_status"),
            num_dependents=num_dep,
            taman_name=val("taman_name"),
            house_number=val("house_number"),
            jalan_aman_serenia=val("jalan_aman_serenia"),
            job_title=val("job_title"),
            employer_name=val("employer_name"),
            employer_address=val("employer_address"),
            employer_phone=val("employer_phone"),
            resident_type=val("resident_type"),
            committee_title=val("committee_title"),
            role=role_val,
            is_active=active,
            is_verified=False,
            status="active" if active else "deactivated",
        )
        db.add(user)
        created += 1

    db.commit()

    return BulkImportResult(created=created, skipped=skipped, errors=errors)
